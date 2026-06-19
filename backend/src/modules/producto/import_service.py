"""Servicio de importación masiva de productos desde Excel QUENDRA."""
import asyncio
import uuid
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.workbook.workbook import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from src.modules.producto.models import Producto
from src.modules.producto import service as producto_service
from src.common.exceptions import BasileException


# Simple in-memory preview cache with TTL
_preview_cache: dict[str, dict] = {}


def _cleanup_expired() -> None:
    """Remove expired entries synchronously."""
    import time
    now = time.time()
    expired = [k for k, v in _preview_cache.items() if v["expires_at"] < now]
    for k in expired:
        _preview_cache.pop(k, None)


def _get_cached(session_id: str):
    _cleanup_expired()
    return _preview_cache.get(session_id)


def _set_cached(session_id: str, preview: "PreviewResult", ttl_seconds: int = 600) -> None:
    _cleanup_expired()
    import time
    _preview_cache[session_id] = {
        "preview": preview,
        "expires_at": time.time() + ttl_seconds,
    }


def _del_cached(session_id: str) -> None:
    _preview_cache.pop(session_id, None)


@dataclass
class PreviewRow:
    row_num: int
    plu: str
    nombre: str
    categoria_nombre: str
    precio_publico: Decimal
    precio_mayorista: Decimal
    costo_por_kilo: Decimal
    stock_actual: Decimal
    stock_minimo: Optional[Decimal]
    errores: list[str] = field(default_factory=list)
    es_valida: bool = True
    es_duplicada: bool = False


@dataclass
class PreviewResult:
    session_id: str
    empresa_id: uuid.UUID
    filas_validas: list[PreviewRow]
    filas_invalidas: list[PreviewRow]
    total_filas: int


def _normalizar_columnas(headers: list[str]) -> list[str]:
    """Normaliza nombres de columnas para mapeo flexible."""
    return [h.strip().lower().replace(" ", "_").replace("-", "_") for h in headers]


def _mapear_columna(normalizado: str) -> Optional[str]:
    """Mapea nombre de columna normalizado al campo interno."""
    mapeo = {
        "plu": "plu",
        "codigo": "plu",
        "codigo_plu": "plu",
        "nombre": "nombre",
        "descripcion": "nombre",
        "categoria": "categoria",
        "categoria_nombre": "categoria",
        "precio_publico": "precio_publico",
        "precio": "precio_publico",
        "precio_venta": "precio_publico",
        "precio_mayorista": "precio_mayorista",
        "costo_kilo": "costo_por_kilo",
        "costo_por_kilo": "costo_por_kilo",
        "costo": "costo_por_kilo",
        "stock_actual": "stock_actual",
        "stock": "stock_actual",
        "stock_minimo": "stock_minimo",
    }
    return mapeo.get(normalizado)


def _parse_decimal(valor, campo: str) -> tuple[Optional[Decimal], Optional[str]]:
    """Parsea un valor a Decimal. Devuelve (valor, error)."""
    if valor is None or str(valor).strip() == "":
        return None, None
    try:
        # Manejar comas como separadores decimales (formato español)
        str_val = str(valor).strip().replace(".", "").replace(",", ".")
        # Pero si el valor ya tiene punto como separador decimal y no tiene coma de miles...
        # Revertimos: si tiene coma, asumimos coma=decimal. Si no, punto=decimal.
        str_val = str(valor).strip()
        if "," in str_val and "." not in str_val:
            str_val = str_val.replace(",", ".")
        elif "," in str_val and "." in str_val:
            # Formato 1.000,50 -> 1000.50
            str_val = str_val.replace(".", "").replace(",", ".")
        d = Decimal(str_val)
        if d < 0:
            return None, f"{campo} no puede ser negativo"
        return d, None
    except (InvalidOperation, ValueError, TypeError):
        return None, f"{campo} debe ser numérico"


def _leer_workbook(data: bytes) -> Workbook:
    """Lee un workbook desde bytes."""
    try:
        return openpyxl.load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:
        raise BasileException(f"No se pudo leer el archivo Excel: {exc}", status_code=400)


def _validar_fila(
    row_data: dict,
    row_num: int,
    categorias_existentes: dict[str, uuid.UUID],
    plus_db: set[str],
    plus_archivo: set[str],
) -> PreviewRow:
    """Valida una fila del Excel y devuelve un PreviewRow."""
    def _safe_str(val) -> str:
        if val is None:
            return ""
        return str(val).strip()

    preview = PreviewRow(
        row_num=row_num,
        plu=_safe_str(row_data.get("plu")),
        nombre=_safe_str(row_data.get("nombre")),
        categoria_nombre=_safe_str(row_data.get("categoria")),
        precio_publico=Decimal("0"),
        precio_mayorista=Decimal("0"),
        costo_por_kilo=Decimal("0"),
        stock_actual=Decimal("0"),
        stock_minimo=None,
    )

    # Validar nombre
    if not preview.nombre:
        preview.errores.append("nombre es obligatorio")
        preview.es_valida = False

    # Validar PLU
    if not preview.plu:
        preview.errores.append("PLU es obligatorio")
        preview.es_valida = False
    elif preview.plu in plus_archivo:
        preview.errores.append("PLU duplicado en archivo")
        preview.es_duplicada = True
        preview.es_valida = False
    elif preview.plu in plus_db:
        preview.errores.append("PLU ya existe en la empresa")
        preview.es_duplicada = True
        preview.es_valida = False

    # Validar categoría
    if preview.categoria_nombre and preview.categoria_nombre.lower() not in {k.lower() for k in categorias_existentes.keys()}:
        preview.errores.append("categoría no encontrada")
        preview.es_valida = False

    # Parsear decimales
    for campo, attr in [
        ("precio_publico", "precio_publico"),
        ("precio_mayorista", "precio_mayorista"),
        ("costo_por_kilo", "costo_por_kilo"),
        ("stock_actual", "stock_actual"),
        ("stock_minimo", "stock_minimo"),
    ]:
        valor_raw = row_data.get(campo)
        valor, error = _parse_decimal(valor_raw, campo)
        if error:
            preview.errores.append(error)
            preview.es_valida = False
        setattr(preview, attr, valor)

    return preview


async def generar_preview(
    db: AsyncSession,
    empresa_id: uuid.UUID,
    file_data: bytes,
) -> PreviewResult:
    """Genera un preview de importación desde un archivo Excel."""
    # Validar tamaño (máximo 5000 filas)
    workbook = await asyncio.to_thread(_leer_workbook, file_data)
    sheet = workbook.active
    if sheet is None:
        raise BasileException("El archivo Excel no tiene hojas", status_code=400)

    row_count = sheet.max_row - 1  # Excluir header
    if row_count > 5000:
        raise BasileException("Máximo 5000 filas por importación", status_code=413)

    # Leer headers
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalizados = _normalizar_columnas(headers)
    mapeo = {idx: _mapear_columna(col) for idx, col in enumerate(normalizados) if _mapear_columna(col)}

    if not mapeo:
        raise BasileException("No se encontraron columnas reconocibles en el Excel", status_code=400)

    # Cargar categorías existentes de la empresa
    from sqlalchemy import select
    from src.modules.producto.models import CategoriaProducto
    cat_result = await db.execute(
        select(CategoriaProducto).where(CategoriaProducto.empresa_id == empresa_id)
    )
    categorias_db = {c.nombre: c.id for c in cat_result.scalars().all()}

    # Cargar PLUs existentes de la empresa
    plu_result = await db.execute(
        select(Producto.plu).where(
            Producto.empresa_id == empresa_id,
            Producto.activo == True,
        )
    )
    plus_db = set(plu_result.scalars().all())

    filas_validas: list[PreviewRow] = []
    filas_invalidas: list[PreviewRow] = []
    plus_archivo: set[str] = set()

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {mapeo[idx]: row[idx] for idx in mapeo if idx < len(row)}
        preview = _validar_fila(row_data, row_idx, categorias_db, plus_db, plus_archivo)

        if preview.plu and not preview.es_duplicada:
            plus_archivo.add(preview.plu)

        if preview.es_valida:
            filas_validas.append(preview)
        else:
            filas_invalidas.append(preview)

    session_id = str(uuid.uuid4())
    result = PreviewResult(
        session_id=session_id,
        empresa_id=empresa_id,
        filas_validas=filas_validas,
        filas_invalidas=filas_invalidas,
        total_filas=row_count,
    )

    # Guardar en cache con TTL 10 minutos
    _set_cached(session_id, result, ttl_seconds=600)

    return result


async def confirmar_importacion(
    db: AsyncSession,
    session_id: str,
    empresa_id: uuid.UUID,
) -> dict:
    """Confirma la importación persistiendo las filas válidas del preview."""
    cached = _get_cached(session_id)
    if not cached:
        raise BasileException("Sesión de preview expirada o inválida", status_code=410)

    preview: PreviewResult = cached["preview"]

    # Verificar que la sesión pertenece a la empresa autenticada
    if preview.empresa_id != empresa_id:
        raise BasileException("Sesión de preview no pertenece a esta empresa", status_code=403)

    from sqlalchemy import select
    from src.modules.producto.models import CategoriaProducto
    cat_result = await db.execute(
        select(CategoriaProducto).where(CategoriaProducto.empresa_id == empresa_id)
    )
    categorias_db = {c.nombre.lower(): c.id for c in cat_result.scalars().all()}

    creados = 0
    errores = 0

    for fila in preview.filas_validas:
        try:
            categoria_id = categorias_db.get(fila.categoria_nombre.lower()) if fila.categoria_nombre else None
            producto = Producto(
                empresa_id=empresa_id,
                plu=fila.plu,
                nombre=fila.nombre,
                categoria_id=categoria_id,
                precio_publico=fila.precio_publico or Decimal("0"),
                precio_mayorista=fila.precio_mayorista or Decimal("0"),
                costo_por_kilo=fila.costo_por_kilo or Decimal("0"),
                stock_actual=fila.stock_actual or Decimal("0"),
                stock_minimo=fila.stock_minimo,
            )
            producto.recalcular_margen()
            db.add(producto)
            creados += 1
        except Exception:
            errores += 1

    await db.commit()

    # Limpiar cache
    _del_cached(session_id)

    return {
        "creados": creados,
        "errores": errores,
        "total_validas": len(preview.filas_validas),
        "total_invalidas": len(preview.filas_invalidas),
    }
