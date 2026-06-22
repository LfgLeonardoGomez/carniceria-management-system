"""End-to-end integration tests for the reporte module.

Uses testcontainers (real PostgreSQL). Tests seed ventas and call the
API through the ASGI test client.

Tasks covered:
  11.1 — seed a Venta with all costs set and one with NULL cost;
          GET /reportes/ventas → both appear with correct ganancia_estimada values
  11.2 — GET /reportes/ventas/exportar?formato=xlsx → valid workbook, ≥2 rows
  11.3 — GET /reportes/ventas/exportar?formato=csv → BOM present, header matches RN-REP-03
  11.4 — GET /reportes/ventas/exportar?formato=pdf → body starts with %PDF
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from decimal import Decimal

import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from src.modules.auth.models import Usuario, Rol
from src.modules.empresa.models import Empresa
from src.modules.producto.models import Producto
from src.modules.venta.models import Venta, DetalleVenta, PagoVenta
from src.core.security import create_access_token, hash_password


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _crear_empresa(db: AsyncSession, nombre: str = "Emp Integ") -> Empresa:
    emp = Empresa(nombre_comercial=nombre, activa=True)
    db.add(emp)
    await db.commit()
    await db.refresh(emp)
    return emp


async def _crear_rol(db: AsyncSession, nombre: str, empresa_id: uuid.UUID | None = None) -> Rol:
    rol = Rol(nombre=nombre, empresa_id=empresa_id)
    db.add(rol)
    await db.commit()
    await db.refresh(rol)
    return rol


async def _crear_usuario(
    db: AsyncSession, email: str, empresa_id: uuid.UUID, rol_id: uuid.UUID
) -> Usuario:
    usuario = Usuario(
        email=email,
        contrasena_hash=hash_password("Pass123!"),
        nombre="Integ",
        apellido="Test",
        rol_id=rol_id,
        activo=True,
        empresa_id=empresa_id,
    )
    db.add(usuario)
    await db.commit()
    await db.refresh(usuario)
    return usuario


async def _crear_producto(db: AsyncSession, empresa_id: uuid.UUID, nombre: str = "Asado") -> Producto:
    prod = Producto(
        empresa_id=empresa_id,
        plu=f"I-{nombre[:3].upper()}-{uuid.uuid4().hex[:4]}",
        nombre=nombre,
        precio_publico=Decimal("1000.00"),
        precio_mayorista=Decimal("900.00"),
        costo_por_kilo=Decimal("500.00"),
        margen=Decimal("0.5000"),
        stock_actual=Decimal("100.000"),
    )
    db.add(prod)
    await db.commit()
    await db.refresh(prod)
    return prod


async def _crear_venta_cobrada(
    db: AsyncSession,
    empresa_id: uuid.UUID,
    producto: Producto,
    costo_unitario: str | None = "500.00",
) -> Venta:
    cantidad = Decimal("2.000")
    precio = Decimal("1000.00")
    importe = (cantidad * precio).quantize(Decimal("0.01"))
    costo_snap = Decimal(costo_unitario) if costo_unitario else None

    detalle = DetalleVenta(
        producto_id=producto.id,
        cantidad_kilos=cantidad,
        precio_unitario=precio,
        importe=importe,
        costo_unitario=costo_snap,
    )
    venta = Venta(
        empresa_id=empresa_id,
        cliente_id=None,
        tipo_cliente_al_momento="publico_general",
        estado="cobrada",
        subtotal=importe,
        descuentos=Decimal("0.00"),
        total=importe,
        fecha=datetime.now(timezone.utc),
    )
    venta.detalles = [detalle]
    db.add(venta)
    await db.commit()
    await db.refresh(venta)

    pago = PagoVenta(venta_id=venta.id, medio_pago="efectivo", importe=importe)
    db.add(pago)
    await db.commit()
    return venta


def _token(usuario_id: uuid.UUID) -> str:
    return create_access_token({"sub": str(usuario_id)})


# ---------------------------------------------------------------------------
# Task 11.1 — ganancia_estimada values correct for both cost-set and NULL-cost ventas
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_integration_ganancia_decimal_and_null(client: AsyncClient, db_session: AsyncSession):
    """Both ganancia_estimada values appear correctly in GET /reportes/ventas."""
    emp = await _crear_empresa(db_session, "Emp 11.1")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"int11-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)
    prod = await _crear_producto(db_session, emp.id, "Asado int")

    # Venta with cost → ganancia = (2*1000) - (2*500) = 1000.00
    await _crear_venta_cobrada(db_session, emp.id, prod, costo_unitario="500.00")
    # Venta with NULL cost → ganancia = null
    await _crear_venta_cobrada(db_session, emp.id, prod, costo_unitario=None)

    response = await client.get(
        "/reportes/ventas",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    data = response.json()

    rows = data["rows"]
    assert len(rows) == 2

    ganancias = {r["ganancia_estimada"] for r in rows}
    assert "1000.00" in ganancias or 1000.0 in {
        float(g) for g in ganancias if g is not None
    }
    assert None in ganancias


# ---------------------------------------------------------------------------
# Task 11.2 — xlsx export produces valid workbook with ≥2 rows (header + 1 data)
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_integration_export_xlsx_valid_workbook(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas/exportar?formato=xlsx → valid workbook with ≥2 rows."""
    import io
    from openpyxl import load_workbook

    emp = await _crear_empresa(db_session, "Emp 11.2")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"int112-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)
    prod = await _crear_producto(db_session, emp.id, "Costilla int")
    await _crear_venta_cobrada(db_session, emp.id, prod)

    response = await client.get(
        "/reportes/ventas/exportar?formato=xlsx",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb["Ventas"]
    assert ws.max_row >= 2, f"Expected at least 2 rows (header + 1 data), got {ws.max_row}"


# ---------------------------------------------------------------------------
# Task 11.3 — CSV export has BOM and correct header
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_integration_export_csv_bom_and_header(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas/exportar?formato=csv → UTF-8 BOM and RN-REP-03 header."""
    import csv, io

    emp = await _crear_empresa(db_session, "Emp 11.3")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"int113-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas/exportar?formato=csv",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200

    content = response.content
    assert content[:3] == b"\xef\xbb\xbf", "Missing UTF-8 BOM"

    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    header = next(reader)

    expected_header = [
        "fecha", "cliente", "productos", "kilos_vendidos",
        "subtotal", "total", "medio_pago", "ganancia_estimada",
    ]
    assert header == expected_header


# ---------------------------------------------------------------------------
# Task 11.4 — PDF export starts with %PDF
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_integration_export_pdf_valid(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas/exportar?formato=pdf → response body starts with %PDF."""
    emp = await _crear_empresa(db_session, "Emp 11.4")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"int114-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas/exportar?formato=pdf",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    assert response.content[:4] == b"%PDF", f"Not a valid PDF header: {response.content[:8]!r}"
