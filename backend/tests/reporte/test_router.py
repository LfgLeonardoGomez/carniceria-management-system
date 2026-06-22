"""Integration tests for the reporte router endpoints.

TDD cycle: RED → GREEN → TRIANGULATE
Uses testcontainers (real PostgreSQL). Every test is isolated via transaction rollback.

Tasks covered:
  6.1 — GET /reportes/ventas returns 200 for administrador; 403 for cajero; 401 without JWT
  6.3 — GET /reportes/ventas/exportar?formato=xlsx → 200 correct Content-Type
         formato=csv → 200 text/csv
         formato=pdf → 200 application/pdf
         formato=docx → 422
  6.5 — Export with date-range filter params are passed through correctly
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone, timedelta
from decimal import Decimal

import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from src.modules.auth.models import Usuario, Rol
from src.modules.empresa.models import Empresa
from src.modules.cliente.models import Cliente
from src.modules.producto.models import Producto
from src.modules.venta.models import Venta, DetalleVenta, PagoVenta
from src.core.security import create_access_token, hash_password


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _crear_empresa(db: AsyncSession, nombre: str = "Emp Router") -> Empresa:
    emp = Empresa(nombre_comercial=nombre, activa=True)
    db.add(emp)
    await db.commit()
    await db.refresh(emp)
    return emp


async def _crear_rol(db: AsyncSession, nombre: str, empresa_id: uuid.UUID | None = None) -> Rol:
    rol = Rol(nombre=nombre, empresa_id=empresa_id)
    db.add(rol)
    await db.commit()
    await db.refresh(rol)
    return rol


async def _crear_usuario(
    db: AsyncSession, email: str, empresa_id: uuid.UUID, rol_id: uuid.UUID
) -> Usuario:
    usuario = Usuario(
        email=email,
        contrasena_hash=hash_password("Pass123!"),
        nombre="Router",
        apellido="Test",
        rol_id=rol_id,
        activo=True,
        empresa_id=empresa_id,
    )
    db.add(usuario)
    await db.commit()
    await db.refresh(usuario)
    return usuario


async def _crear_producto(
    db: AsyncSession, empresa_id: uuid.UUID, nombre: str = "Asado"
) -> Producto:
    prod = Producto(
        empresa_id=empresa_id,
        plu=f"R-{nombre[:3].upper()}-{uuid.uuid4().hex[:4]}",
        nombre=nombre,
        precio_publico=Decimal("1000.00"),
        precio_mayorista=Decimal("900.00"),
        costo_por_kilo=Decimal("500.00"),
        margen=Decimal("0.5000"),
        stock_actual=Decimal("100.000"),
    )
    db.add(prod)
    await db.commit()
    await db.refresh(prod)
    return prod


async def _crear_venta_cobrada(
    db: AsyncSession,
    empresa_id: uuid.UUID,
    producto: Producto,
    fecha: datetime | None = None,
    costo_unitario: str | None = "500.00",
) -> Venta:
    if fecha is None:
        fecha = datetime.now(timezone.utc)
    cantidad = Decimal("2.000")
    precio = Decimal("1000.00")
    importe = (cantidad * precio).quantize(Decimal("0.01"))
    costo_snap = Decimal(costo_unitario) if costo_unitario else None

    detalle = DetalleVenta(
        producto_id=producto.id,
        cantidad_kilos=cantidad,
        precio_unitario=precio,
        importe=importe,
        costo_unitario=costo_snap,
    )
    venta = Venta(
        empresa_id=empresa_id,
        cliente_id=None,
        tipo_cliente_al_momento="publico_general",
        estado="cobrada",
        subtotal=importe,
        descuentos=Decimal("0.00"),
        total=importe,
        fecha=fecha,
    )
    venta.detalles = [detalle]
    db.add(venta)
    await db.commit()
    await db.refresh(venta)

    pago = PagoVenta(venta_id=venta.id, medio_pago="efectivo", importe=importe)
    db.add(pago)
    await db.commit()
    return venta


def _token(usuario_id: uuid.UUID) -> str:
    return create_access_token({"sub": str(usuario_id)})


# ---------------------------------------------------------------------------
# Task 6.1 — Access control
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_get_reportes_ventas_admin_200(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas returns 200 for administrador role."""
    emp = await _crear_empresa(db_session, "Emp admin 200")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"admin-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200


@pytest.mark.asyncio
async def test_get_reportes_ventas_cajero_403(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas returns 403 for cajero role."""
    emp = await _crear_empresa(db_session, "Emp cajero 403")
    rol = await _crear_rol(db_session, "cajero", emp.id)
    usuario = await _crear_usuario(db_session, f"cajero-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 403


@pytest.mark.asyncio
async def test_get_reportes_ventas_no_jwt_401(client: AsyncClient):
    """GET /reportes/ventas returns 401 without JWT."""
    response = await client.get("/reportes/ventas")
    assert response.status_code == 401


# ---------------------------------------------------------------------------
# Task 6.3 — Export format endpoints
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_exportar_xlsx_200(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas/exportar?formato=xlsx returns 200 with correct Content-Type."""
    emp = await _crear_empresa(db_session, "Emp xlsx export")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"xlsx-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)
    prod = await _crear_producto(db_session, emp.id, "Asado xlsx")
    await _crear_venta_cobrada(db_session, emp.id, prod)

    response = await client.get(
        "/reportes/ventas/exportar?formato=xlsx",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers.get("content-type", "")
    assert len(response.content) > 0


@pytest.mark.asyncio
async def test_exportar_csv_200(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas/exportar?formato=csv returns 200 text/csv."""
    emp = await _crear_empresa(db_session, "Emp csv export")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"csv-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas/exportar?formato=csv",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    assert "text/csv" in response.headers.get("content-type", "")


@pytest.mark.asyncio
async def test_exportar_pdf_200(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas/exportar?formato=pdf returns 200 application/pdf."""
    emp = await _crear_empresa(db_session, "Emp pdf export")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"pdf-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas/exportar?formato=pdf",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    assert "application/pdf" in response.headers.get("content-type", "")


@pytest.mark.asyncio
async def test_exportar_docx_422(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas/exportar?formato=docx returns 422."""
    emp = await _crear_empresa(db_session, "Emp docx 422")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"docx-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas/exportar?formato=docx",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 422


# ---------------------------------------------------------------------------
# Task 6.5 — TRIANGULATE: date-range filter params pass through to export
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_exportar_date_range_filter_pass_through(client: AsyncClient, db_session: AsyncSession):
    """Export with date-range filter returns only rows matching those dates."""
    emp = await _crear_empresa(db_session, "Emp export date filter")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"drf-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)
    prod = await _crear_producto(db_session, emp.id, "Bife date")

    now = datetime.now(timezone.utc)
    # Create one venta yesterday, one two_days_ago
    yesterday = now - timedelta(days=1)
    await _crear_venta_cobrada(db_session, emp.id, prod, fecha=yesterday)

    # Export only from last 12 hours (should find 0 rows — file is still valid xlsx)
    fecha_desde = (now - timedelta(hours=12)).strftime("%Y-%m-%dT%H:%M:%SZ")
    fecha_hasta = now.strftime("%Y-%m-%dT%H:%M:%SZ")

    response = await client.get(
        "/reportes/ventas/exportar",
        params={"formato": "xlsx", "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta},
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200

    # Workbook must be valid even with zero data rows
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb["Ventas"]
    # Only header row, no data rows
    assert ws.max_row == 1


# ---------------------------------------------------------------------------
# Task 6.6 — Content-Disposition filename includes date range (spec requirement)
# Fallback rule:
#   both dates present  → ventas-<fecha_desde_date>-<fecha_hasta_date>.<fmt>
#   both dates absent   → ventas.<fmt>
#   only desde present  → ventas-<fecha_desde_date>-all.<fmt>
#   only hasta present  → ventas-all-<fecha_hasta_date>.<fmt>
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_exportar_filename_with_both_dates(client: AsyncClient, db_session: AsyncSession):
    """Content-Disposition filename includes fecha_desde and fecha_hasta when both are provided."""
    emp = await _crear_empresa(db_session, "Emp filename dates")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"fname-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas/exportar",
        params={
            "formato": "xlsx",
            "fecha_desde": "2026-06-01T00:00:00Z",
            "fecha_hasta": "2026-06-22T23:59:59Z",
        },
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    cd = response.headers.get("content-disposition", "")
    assert "ventas-2026-06-01-2026-06-22.xlsx" in cd


@pytest.mark.asyncio
async def test_exportar_filename_no_dates(client: AsyncClient, db_session: AsyncSession):
    """Content-Disposition filename falls back to ventas.<fmt> when no dates are provided."""
    emp = await _crear_empresa(db_session, "Emp filename no dates")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"fnnd-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas/exportar",
        params={"formato": "csv"},
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    cd = response.headers.get("content-disposition", "")
    assert "ventas.csv" in cd
    # Must NOT contain a date pattern like ventas-20
    assert "ventas-20" not in cd


@pytest.mark.asyncio
async def test_exportar_filename_only_fecha_desde(client: AsyncClient, db_session: AsyncSession):
    """Content-Disposition filename uses fecha_desde with 'all' token when only desde is provided."""
    emp = await _crear_empresa(db_session, "Emp filename desde only")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"fndo-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)

    response = await client.get(
        "/reportes/ventas/exportar",
        params={"formato": "pdf", "fecha_desde": "2026-01-01T00:00:00Z"},
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    cd = response.headers.get("content-disposition", "")
    assert "ventas-2026-01-01-all.pdf" in cd


@pytest.mark.asyncio
async def test_get_reportes_ventas_returns_paginated_json(client: AsyncClient, db_session: AsyncSession):
    """GET /reportes/ventas returns paginated JSON with rows/total/skip/limit."""
    emp = await _crear_empresa(db_session, "Emp paginated")
    rol = await _crear_rol(db_session, "administrador", emp.id)
    usuario = await _crear_usuario(db_session, f"pag-{uuid.uuid4().hex[:4]}@test.com", emp.id, rol.id)
    prod = await _crear_producto(db_session, emp.id, "Vacío paginated")
    await _crear_venta_cobrada(db_session, emp.id, prod)

    response = await client.get(
        "/reportes/ventas",
        headers={"Authorization": f"Bearer {_token(usuario.id)}"},
    )
    assert response.status_code == 200
    data = response.json()
    assert "rows" in data
    assert "total" in data
    assert "skip" in data
    assert "limit" in data
    assert data["total"] >= 1
