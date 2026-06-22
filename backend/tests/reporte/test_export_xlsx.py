"""Tests for reporte.service.generar_xlsx.

TDD cycle: RED → GREEN → TRIANGULATE
Tasks:
  3.1 — generar_xlsx returns valid bytes decodable as workbook, sheet "Ventas",
         correct header row, data rows matching input, numeric types, blank for null ganancia
  3.3 — accented characters in product/client names preserved (no encoding error)
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from decimal import Decimal

import pytest


def _make_row(
    cliente_nombre: str = "Juan Perez",
    productos: str = "Asado",
    total_kilos: str = "2.000",
    subtotal: str = "2000.00",
    total: str = "2000.00",
    medios_pago: str = "efectivo",
    ganancia_estimada: str | None = "500.00",
) -> "VentaReporteRow":
    from src.modules.reporte.schemas import VentaReporteRow
    return VentaReporteRow(
        venta_id=uuid.uuid4(),
        fecha=datetime(2024, 6, 1, 10, 0, 0, tzinfo=timezone.utc),
        cliente_nombre=cliente_nombre,
        productos=productos,
        total_kilos=Decimal(total_kilos),
        subtotal=Decimal(subtotal),
        total=Decimal(total),
        medios_pago=medios_pago,
        ganancia_estimada=Decimal(ganancia_estimada) if ganancia_estimada is not None else None,
    )


# ---------------------------------------------------------------------------
# Task 3.1 — RED: valid xlsx bytes, sheet name, header, data rows, numeric types
# ---------------------------------------------------------------------------

def test_generar_xlsx_returns_valid_bytes():
    """generar_xlsx returns non-empty bytes that parse as a workbook."""
    from openpyxl import load_workbook
    from src.modules.reporte.service import generar_xlsx

    row = _make_row()
    result = generar_xlsx([row])

    assert isinstance(result, bytes)
    assert len(result) > 0

    import io
    wb = load_workbook(io.BytesIO(result))
    assert "Ventas" in wb.sheetnames


def test_generar_xlsx_header_row():
    """The first row must be the RN-REP-03 header."""
    from openpyxl import load_workbook
    from src.modules.reporte.service import generar_xlsx
    import io

    result = generar_xlsx([_make_row()])
    ws = load_workbook(io.BytesIO(result))["Ventas"]

    expected_headers = [
        "fecha", "cliente", "productos", "kilos_vendidos",
        "subtotal", "total", "medio_pago", "ganancia_estimada",
    ]
    actual = [cell.value for cell in ws[1]]
    assert actual == expected_headers


def test_generar_xlsx_data_row_values():
    """Data row values match the input row."""
    from openpyxl import load_workbook
    from src.modules.reporte.service import generar_xlsx
    import io

    row = _make_row(
        cliente_nombre="Maria Lopez",
        productos="Bife de lomo",
        total_kilos="3.500",
        subtotal="3500.00",
        total="3500.00",
        medios_pago="tarjeta",
        ganancia_estimada="700.00",
    )
    result = generar_xlsx([row])
    ws = load_workbook(io.BytesIO(result))["Ventas"]

    data_row = [cell.value for cell in ws[2]]
    assert data_row[1] == "Maria Lopez"     # cliente
    assert data_row[2] == "Bife de lomo"    # productos
    assert abs(data_row[3] - 3.5) < 0.0001  # kilos — numeric
    assert abs(data_row[4] - 3500.0) < 0.01  # subtotal — numeric
    assert abs(data_row[7] - 700.0) < 0.01  # ganancia — numeric


def test_generar_xlsx_null_ganancia_is_blank_cell():
    """Null ganancia_estimada produces a blank cell (None), not the string 'None'."""
    from openpyxl import load_workbook
    from src.modules.reporte.service import generar_xlsx
    import io

    row = _make_row(ganancia_estimada=None)
    result = generar_xlsx([row])
    ws = load_workbook(io.BytesIO(result))["Ventas"]

    ganancia_cell = ws.cell(row=2, column=8)
    assert ganancia_cell.value is None, f"Expected None, got {ganancia_cell.value!r}"


# ---------------------------------------------------------------------------
# Task 3.3 — TRIANGULATE: accented characters preserved
# ---------------------------------------------------------------------------

def test_generar_xlsx_accented_characters_preserved():
    """Accented characters in product/client names don't raise an error and are preserved."""
    from openpyxl import load_workbook
    from src.modules.reporte.service import generar_xlsx
    import io

    row = _make_row(
        cliente_nombre="Público General",
        productos="Costilla de cerdo, Ñoqui",
    )
    # Must not raise
    result = generar_xlsx([row])

    ws = load_workbook(io.BytesIO(result))["Ventas"]
    assert ws.cell(row=2, column=2).value == "Público General"
    assert ws.cell(row=2, column=3).value == "Costilla de cerdo, Ñoqui"
