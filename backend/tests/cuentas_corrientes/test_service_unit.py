"""Unit tests for cuenta_corriente service (pure logic, no DB).

Tests the export generators and verifies the logic can be imported.
These run without Docker/testcontainers.

Integration tests for registrar_pago/obtener_historial live in test_service_pago.py
and require Docker+testcontainers.
"""
from __future__ import annotations

import sys
import uuid
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "backend" / "src"))


# ---------------------------------------------------------------------------
# Service imports
# ---------------------------------------------------------------------------

class TestServiceImports:
    def test_service_module_importable(self):
        from src.modules.cuenta_corriente import service
        assert hasattr(service, "registrar_pago")
        assert hasattr(service, "obtener_historial")
        assert hasattr(service, "obtener_estado_cuenta")
        assert hasattr(service, "generar_xlsx")
        assert hasattr(service, "generar_csv")
        assert hasattr(service, "generar_pdf")

    def test_schemas_module_importable(self):
        from src.modules.cuenta_corriente.schemas import (
            PagoCreate, MovimientoCCResponse, PagoResponse, HistorialCCResponse, EstadoCuentaResponse
        )
        assert PagoCreate is not None


# ---------------------------------------------------------------------------
# Export generators — pure unit tests (no DB)
# ---------------------------------------------------------------------------

def _make_estado(n_movimientos: int = 2):
    """Build an EstadoCuentaResponse in-memory for export testing."""
    from src.modules.cuenta_corriente.schemas import EstadoCuentaResponse, MovimientoCCResponse

    cliente_id = uuid.uuid4()
    movs = []
    running = Decimal("0.00")
    for i in range(n_movimientos):
        importe = Decimal(str((i + 1) * 100))
        running = (running + importe).quantize(Decimal("0.01"))
        movs.append(
            MovimientoCCResponse(
                id=uuid.uuid4(),
                tipo="deuda",
                importe=importe,
                saldo_resultante=running,
                venta_id=None,
                fecha=datetime.utcnow(),
            )
        )

    return EstadoCuentaResponse(
        cliente_id=cliente_id,
        cliente_nombre="Juan Perez",
        saldo_actual=running,
        movimientos=movs,
    )


class TestGenerarXlsx:
    def test_returns_non_empty_bytes(self):
        from src.modules.cuenta_corriente.service import generar_xlsx
        estado = _make_estado(2)
        result = generar_xlsx(estado)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_workbook(self):
        """The bytes should parse as a valid openpyxl workbook."""
        import io
        from openpyxl import load_workbook
        from src.modules.cuenta_corriente.service import generar_xlsx

        estado = _make_estado(3)
        result = generar_xlsx(estado)
        wb = load_workbook(io.BytesIO(result))
        assert "Estado de Cuenta" in wb.sheetnames

    def test_workbook_contains_movement_rows(self):
        """Workbook must have at least as many rows as movements (plus header rows)."""
        import io
        from openpyxl import load_workbook
        from src.modules.cuenta_corriente.service import generar_xlsx

        n = 3
        estado = _make_estado(n)
        result = generar_xlsx(estado)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # 3 header rows + 1 column header row + n movement rows + 1 footer = n + 5
        assert ws.max_row >= n + 2


class TestGenerarCsv:
    def test_returns_non_empty_bytes(self):
        from src.modules.cuenta_corriente.service import generar_csv
        estado = _make_estado(2)
        result = generar_csv(estado)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_contains_movement_data(self):
        """CSV bytes should contain the word 'deuda' from the tipo column."""
        from src.modules.cuenta_corriente.service import generar_csv
        estado = _make_estado(2)
        result = generar_csv(estado)
        # Decode removing BOM
        text = result.decode("utf-8-sig")
        assert "deuda" in text

    def test_contains_cliente_nombre(self):
        from src.modules.cuenta_corriente.service import generar_csv
        estado = _make_estado(1)
        result = generar_csv(estado)
        text = result.decode("utf-8-sig")
        assert "Juan Perez" in text


class TestGenerarPdf:
    def test_returns_non_empty_bytes(self):
        from src.modules.cuenta_corriente.service import generar_pdf
        estado = _make_estado(2)
        result = generar_pdf(estado)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_starts_with_pdf_magic_bytes(self):
        """%PDF is the PDF file signature."""
        from src.modules.cuenta_corriente.service import generar_pdf
        estado = _make_estado(2)
        result = generar_pdf(estado)
        assert result[:4] == b"%PDF"

    def test_empty_movements_produces_valid_pdf(self):
        """Edge case: customer with no movements should still produce a valid PDF."""
        from src.modules.cuenta_corriente.service import generar_pdf
        estado = _make_estado(0)
        result = generar_pdf(estado)
        assert result[:4] == b"%PDF"
